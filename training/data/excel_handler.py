from openpyxl import load_workbook
import numpy as np

from shared.config import DATA_RAW_FILE, LABELS, LABELS_DIR, DATA_PARSED_FILE

# Устаревшее

# def load_excel_data(path=None) -> tuple[list[str], dict[str, np.ndarray]]:
#     """Подгружаем таблицу, парсим"""
#     print(f"Грузим эксель таблицу из {path or DATA_RAW_FILE}")
#     workbook = load_workbook(path or DATA_RAW_FILE)
#     sheet = workbook.active
#     return parse_excel_rows(sheet)


# def get_headers(sheet) -> list[str]:
#     """Получаем заголовки столбцов"""
#     print("Получаем заголовки")
#     return [str(cell).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=sheet.max_column, values_only=True))]


# def parse_excel_rows(sheet) -> tuple[list[str], dict[str, np.ndarray]]:
#     """Парсим строки"""
#     print("Парсим данные")
#     headers = get_headers(sheet) # Заголовки
    
#     # Словарик полезных заголовков и индексов, полезные заголовки - которые есть и в файле и в конфиге
#     header_indices = {header: i + 1 for i, header in enumerate(headers) if header in LABELS} 
#     print(f"Столбцы к обработке: {header_indices.keys()}")
    
#     # словарик меток
#     phrases = []
#     labels = {f"{header}": [] for header in header_indices.keys()}

#     # Парсим полезные столбцы
#     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
#         phrases.append(str(row[0]).strip())
#         for header, i in header_indices.items():
#             #print(f"Смотрим столбец {header}, индекс {i}")
#             labels[header].append(convert_label(header, str(row[i]).strip()))
            
#     labels_numpy = {header: np.array(label) for header, label in labels.items()}
#     return phrases, labels_numpy

    
# def convert_label(name: str, value: str) -> int:
#     """Конвертация с настройкой в конфиге, чтобы сюда не лезть"""
#     if(value == 'Null'): return value
#     return LABELS[name].index(value)



def new__parse_excel(path : str | None = None) -> dict[str, np.ndarray]:
    """Обработка сырой Excel разметки"""
    wb = load_workbook(path or DATA_RAW_FILE)
    sheet = wb.active
    
    print("Парсим данные")
    
    result = {}
    for i, col in enumerate(sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)):
        if i == 0 or col[0] in LABELS:   # первый столбец - сообщения, всегда идет в словарь, остальное только если есть в конфиге
            header = col[0]              # привязка к конфигу позволяет в целом свободно редачить файл, добавлять новую инфу, которая при этом не улетит в машинки и не создаст модели из мусора
            labels = np.array(col[1:])
            result[header] = new__convert_labels(header, labels) if i != 0 else labels
    
    return result
        
def new__load_excel(path: str | None = None) -> dict[str, np.ndarray]:
    """Загрузка уже обработанных данных из файла"""
    return dict(np.load(f"{path}.npz"))

def new__save_excel(data: dict[str, np.ndarray]) -> None:
    """Сохранение обработанного Excel файла"""
    np.savez(DATA_PARSED_FILE, **data)
       
def new__convert_labels(name: str, labels : np.ndarray) -> np.ndarray:
    """Конвертация меток на базе конфига"""
    return [LABELS[name].index(label) if label != 'Null' else -1 for label in labels]

def new__get_excel(path_parsed: str | None = None, path_excel: str | None = None) -> dict[str, np.ndarray]:
    """Единый метод загрузки Excel данных"""
    try:
        return new__load_excel(path_parsed or DATA_PARSED_FILE)
    except Exception as e:
        print(f"Файл обработанных данных не найден, попытка обработать сырые данные")
        result = new__parse_excel(path_excel or DATA_RAW_FILE)
        new__save_excel(result)
        return result
        
    
"""Архитектура не до конца ясна для методов ниже, пока не используются"""

def save_labels(name: str, labels: list[str]) -> None:
    """Сохранение меток в разные файлы по имени (WIP)"""
    LABELS_DIR.mkdir(parents=True, exist_ok=True)
    np.save(LABELS_DIR / f"{name}.npy", np.array(labels))
    
    
def load_labels(name: str) -> np.ndarray:
    """Загрузка меток по имени (WIP)"""
    data = np.load(LABELS_DIR / f"{name}.npy")
    return data
    