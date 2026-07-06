from openpyxl import load_workbook
from config import DATA_FILE, EXCEL_RANGE, PHRASE_COLUMN, LABEL_COLUMN, LABEL_POSITIVE


def load_excel_data() -> tuple[list[str], list[int]]:
    """Загружает фразы и метки из Excel файла."""
    workbook = load_workbook(DATA_FILE)
    sheet = workbook.active
    return parse_excel_rows(sheet[EXCEL_RANGE])


def parse_excel_rows(rows) -> tuple[list[str], list[int]]:
    """Парсит строки Excel в списки фраз и меток."""
    data_rows = list(rows)[1:]  # Пропускаем заголовок
    phrases = [str(row[PHRASE_COLUMN].value) for row in data_rows]
    labels = [convert_label(row[LABEL_COLUMN].value) for row in data_rows]
    return phrases, labels


def convert_label(value: str) -> int:
    """Конвертирует текстовую метку в числовую."""
    return 1 if value == LABEL_POSITIVE else 0


def save_excel_data() -> None:
    """(WIP) Сохраняет данные в Excel файл. Нужен будет для правок за ботом и дообучения"""
    workbook = load_workbook(DATA_FILE)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=PHRASE_COLUMN + 1, max_col=LABEL_COLUMN + 1):
        phrase_cell, label_cell = row
        # Здесь можно добавить код для обновления значений ячеек
        # Например, phrase_cell.value = "новая фраза"
        # label_cell.value = "новая метка"
    workbook.save(DATA_FILE)